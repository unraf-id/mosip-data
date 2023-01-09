import psycopg2
import sys
import argparse
import json
import pandas as pd
from api import *
from utils import *
import openpyxl
from openpyxl.styles import Font
from sqlalchemy import create_engine


def main():
    args, parser =  args_parse()

    tables = args.user_tab_list.split(",")

    engine = create_engine('postgresql://%s:%s@%s:%s/%s' % (args.db_user, args.db_pwd, args.db_host, args.db_port, args.user_db_schema))
    connection = engine.raw_connection()

    if len(tables) > 1:
        for table in tables:
            export_to_excel(connection, table, args.xls_folder)
    else:
        query_string_for_tables='select distinct table_name from information_schema.tables where table_schema not in (\'pg_catalog\', \'information_schema\')'
        tables_list=engine.execute(query_string_for_tables)
        for rowno, row in enumerate(tables_list, start = 2):
       	    for colno, cell_value in enumerate(row, start = 1):
       	        export_to_excel(connection, cell_value, args.xls_folder)

#    table_order = get_order_from_list(os.path.join(sys.path[0], 'table_order'))

#    init_logger('full', 'a', './out.log', level=logging.INFO)  # Append mode

#    upload_xlsx(files, table_order, args.user, args.db_user, args.db_pwd, args.db_host, args.db_port)


def args_parse():
    parser = argparse.ArgumentParser()
    parser.add_argument('db_host', help='db host name/ip address')
    parser.add_argument('db_pwd', help='db host name/ip address')
    parser.add_argument('xls_folder', help='directory containing all the tables in xlsx format')
    parser.add_argument('--db_user', help='Db supseradmin user role for masterdb. Default: postgres', type=str,
                        default='postgres')
    parser.add_argument('--db_port', help='db port. Default is 5432', type=int, default=5432)
    parser.add_argument('user_db_schema', help='DB schema which required to download tables in xlsx format')
    parser.add_argument('user_tab_list', help='List of Tables which required to download')

    args = parser.parse_args()
    return args, parser


def export_to_excel(connection, table, xls_folder):
    query_string='SELECT * FROM %s' % (table)
    cursor=connection.cursor()
    cursor.execute(query_string)
    data=cursor.fetchall()
    column_names = cursor.description
    cursor.close()

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold = True)

    # Spreadsheet row and column indexes start at 1
    # so we use "start = 1" in enumerate so
    # we don't need to add 1 to the indexes.
    colno=1
    for i in column_names:
        print(i)
        sheet.cell(row = 1, column = colno).value = i[0]
        colno=colno+1

    # This time we use "start = 2" to skip the heading row.
    for rowno, row in enumerate(data, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            sheet.cell(row = rowno, column = colno).value = cell_value

    if not os.path.isdir(xls_folder):
        os.makedirs(xls_folder)

    filepath = xls_folder + '/' + table + '.xlsx'
    wb.save(filepath)


if __name__=="__main__":
    main()
